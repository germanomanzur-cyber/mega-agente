#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ─────────────────────────────────────────────────────────────────────────────
# REPORTE SEMANAL DE LEADS DE NICO  (reemplaza al agente de Claude, cero tokens LLM)
# ─────────────────────────────────────────────────────────────────────────────
# Lee /chats.json, toma las conversaciones con actividad en los ultimos 7 dias,
# clasifica cada lead (caliente/tibio/frio), detecta que busca, codigos [REF] y
# si Nico derivo a German, y arma un Excel con 2 hojas:
#   - "Leads":   una fila por conversacion
#   - "Resumen": totales por plataforma / interes / derivados
#
# Toda la clasificacion corre en codigo (sin LLM).
#
# USO:   REPORT_TOKEN=mega-radar-2024 python3 tasks/reporte-leads.py
# ENV:
#   REPORT_URL    (default https://mega-agente-production.up.railway.app)
#   REPORT_TOKEN  (default 'mega-radar-2024')
#   GERMAN_PHONE  (default '5493424287842')  -> se excluye del reporte
#   OUT_DIR       (default '.')              -> carpeta donde se guarda el .xlsx
#   DIAS          (default '7')              -> ventana de dias hacia atras
# ─────────────────────────────────────────────────────────────────────────────
import os
import re
import sys
import json
import urllib.request
import urllib.parse
from datetime import datetime, timedelta, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPORT_URL = os.environ.get("REPORT_URL", "https://mega-agente-production.up.railway.app").rstrip("/")
REPORT_TOKEN = os.environ.get("REPORT_TOKEN", "mega-radar-2024")
GERMAN_PHONE = os.environ.get("GERMAN_PHONE", "5493424287842")
OUT_DIR = os.environ.get("OUT_DIR", ".")
DIAS = int(os.environ.get("DIAS", "7"))

# zona horaria Argentina (-03:00)
AR_TZ = timezone(timedelta(hours=-3))

# ── clasificacion de interes (misma heuristica que el monitor diario) ────────
KW_CALIENTE = [
    "compr", "señ", "sena", "reserv", "visit", "ver la propiedad", "ver la casa", "ir a ver",
    "cuando puedo ver", "cuándo puedo ver", "coordin", "efectivo", "credito aprob", "crédito aprob",
    "hipotecario aprob", "quiero", "me la quedo", "hago la oferta", "ofrezco", "disponible para mudar",
    "llamame", "llámame", "mi numero", "mi número", "pasame", "agend",
]
KW_TIBIO = [
    "info", "informacion", "información", "precio", "cuanto", "cuánto", "valor", "disponible",
    "consult", "interesa", "metros", "ambientes", "dormitorios", "cochera", "financ", "credito", "crédito",
    "permuta", "expensas", "fotos", "ubicacion", "ubicación", "zona",
]
ZONAS = [
    "centro", "candioti", "guadalupe", "norte", "sur", "parque", "colastine", "colastiné", "sauce viejo",
    "santo tome", "santo tomé", "rincon", "rincón", "recreo", "arroyo leyes", "arroyo aguiar", "parana",
    "paraná", "oro verde", "monte vera", "bajada grande", "villa california", "san benito", "colonia avellaneda",
]
TIPOS = ["casa", "departamento", "depto", "ph", "terreno", "lote", "quinta", "local", "oficina",
         "galpon", "galpón", "cochera", "duplex", "dúplex"]


def clasificar(user_text):
    t = user_text.lower()
    if any(k in t for k in KW_CALIENTE):
        return "caliente"
    if any(k in t for k in KW_TIBIO):
        return "tibio"
    return "frio"


def extraer(user_texts):
    t = " ".join(user_texts).lower()
    zona = next((z for z in ZONAS if z in t), "")
    tipo = next((x for x in TIPOS if x in t), "")
    pres = ""
    m_usd = re.search(r"(u\$s|usd|us\$|d[oó]lares)\s*([\d.,]+)", t)
    m_mil = re.search(r"([\d.,]+)\s*(mil|k|lucas|palos|millones?)", t)
    m_num = re.search(r"\$\s*([\d.][\d.,]{3,})", t)
    if m_usd:
        pres = f"USD {m_usd.group(2)}"
    elif m_mil:
        pres = f"{m_mil.group(1)} {m_mil.group(2)}"
    elif m_num:
        pres = f"${m_num.group(1)}"
    partes = [p for p in (tipo, zona, pres) if p]
    return " / ".join(partes)


def refs(nico_texts):
    txt = " ".join(nico_texts)
    out = []
    seen = set()
    def add(c):
        c = c.strip()
        if c and c not in seen:
            seen.add(c)
            out.append(c)
    for m in re.finditer(r"\[([A-Z0-9][A-Z0-9\-\s]{1,20})\]", txt):
        add(re.sub(r"^REF[:\s.]+", "", m.group(1).strip(), flags=re.I).strip())
    for m in re.finditer(r"\b(M[A-Z]{2}\d{5,})\b", txt):
        add(m.group(1))
    for m in re.finditer(r"\bref[:\s.]+([A-Z0-9\-]{3,})", txt, flags=re.I):
        add(m.group(1))
    return out


def derivado(nico_texts):
    txt = " ".join(nico_texts).lower()
    if "5493424287842" in txt or "wa.me/549342" in txt:
        return True
    return bool(re.search(r"\bgerm[aá]n\b", txt) and re.search(r"(contact|deriv|escrib|llam|coordina|pasa el)", txt))


def parse_at(s):
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except Exception:
        return None


def get_json(url):
    req = urllib.request.Request(url, headers={"User-Agent": "nico-leads"})
    with urllib.request.urlopen(req, timeout=30) as r:
        return json.loads(r.read().decode("utf-8"))


def main():
    ahora = datetime.now(AR_TZ)
    desde = ahora - timedelta(days=DIAS)
    fecha_archivo = ahora.strftime("%Y-%m-%d")
    fecha_pretty = ahora.strftime("%d/%m/%Y")

    url = f"{REPORT_URL}/chats.json?token={urllib.parse.quote(REPORT_TOKEN)}"
    try:
        chats = get_json(url)
    except Exception as e:
        print(f"[leads] ERROR leyendo /chats.json: {e}", file=sys.stderr)
        sys.exit(1)

    leads = []
    for key, c in chats.items():
        if not isinstance(c, dict):
            continue
        msgs = c.get("messages")
        if not isinstance(msgs, list) or not msgs:
            continue
        user_id = str(c.get("userId") or key)
        if GERMAN_PHONE in user_id:
            continue

        # actividad dentro de la ventana
        recientes = []
        for m in msgs:
            at = parse_at(m.get("at", "")) if isinstance(m, dict) else None
            if at and at.astimezone(AR_TZ) >= desde:
                recientes.append(m)
        if not recientes:
            continue

        user_texts = [m.get("text", "") for m in msgs if m.get("role") != "nico"]
        nico_texts = [m.get("text", "") for m in msgs if m.get("role") == "nico"]
        nivel = clasificar(" ".join(user_texts))
        busca = extraer(user_texts) or "consulta general"
        codigos = refs(nico_texts)
        fue_derivado = derivado(nico_texts)

        plat = c.get("channel") or "?"
        nombre = c.get("waName") or c.get("name") or user_id

        ats = [parse_at(m.get("at", "")) for m in msgs if m.get("at")]
        ats = [a.astimezone(AR_TZ) for a in ats if a]
        primer = min(ats).strftime("%d/%m/%Y %H:%M") if ats else ""
        ultimo = max(ats).strftime("%d/%m/%Y %H:%M") if ats else ""

        leads.append({
            "nombre": nombre,
            "plataforma": plat,
            "busca": busca,
            "interes": nivel,
            "codigos": ", ".join(codigos),
            "derivado": "Sí" if fue_derivado else "No",
            "mensajes": len(msgs),
            "primer": primer,
            "ultimo": ultimo,
            "userId": user_id,
        })

    # ordenar: calientes primero, luego tibios, luego frios; dentro por ultima actividad desc
    orden = {"caliente": 0, "tibio": 1, "frio": 2}
    leads.sort(key=lambda x: (orden.get(x["interes"], 9), x["ultimo"]), reverse=False)
    leads.sort(key=lambda x: orden.get(x["interes"], 9))

    # ── armar Excel ──────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    headers = ["Nombre", "Plataforma", "Qué busca", "Interés", "Códigos REF",
               "Derivado a Germán", "Mensajes", "Primer contacto", "Última actividad", "Contacto/ID"]
    ws.append(headers)

    # estilos
    head_fill = PatternFill("solid", fgColor="1F4E78")
    head_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_cal = PatternFill("solid", fgColor="FCE4E4")  # rojo claro
    fill_tib = PatternFill("solid", fgColor="FFF2CC")  # amarillo claro
    fill_fri = PatternFill("solid", fgColor="EDEDED")  # gris claro

    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = center
        cell.border = border

    interes_label = {"caliente": "🔴 Caliente", "tibio": "🟡 Tibio", "frio": "⚪ Frío"}
    for ld in leads:
        ws.append([
            ld["nombre"], ld["plataforma"], ld["busca"], interes_label.get(ld["interes"], ld["interes"]),
            ld["codigos"], ld["derivado"], ld["mensajes"], ld["primer"], ld["ultimo"], ld["userId"],
        ])
        r = ws.max_row
        fill = fill_cal if ld["interes"] == "caliente" else fill_tib if ld["interes"] == "tibio" else fill_fri
        for col in range(1, len(headers) + 1):
            cl = ws.cell(row=r, column=col)
            cl.border = border
            if col == 4:
                cl.fill = fill
                cl.alignment = center
            if col in (2, 6, 7):
                cl.alignment = center

    # anchos
    widths = [22, 11, 34, 13, 18, 16, 10, 17, 17, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    # ── hoja Resumen ─────────────────────────────────────────────────────────
    rs = wb.create_sheet("Resumen")
    total = len(leads)
    n_cal = sum(1 for l in leads if l["interes"] == "caliente")
    n_tib = sum(1 for l in leads if l["interes"] == "tibio")
    n_fri = sum(1 for l in leads if l["interes"] == "frio")
    n_der = sum(1 for l in leads if l["derivado"] == "Sí")
    por_plat = {}
    for l in leads:
        por_plat[l["plataforma"]] = por_plat.get(l["plataforma"], 0) + 1

    rs.append(["REPORTE SEMANAL DE LEADS — NICO"])
    rs["A1"].font = Font(bold=True, size=14, color="1F4E78")
    rs.append([f"Período: últimos {DIAS} días (al {fecha_pretty})"])
    rs.append([])
    rs.append(["Métrica", "Valor"])
    for cell in (rs["A4"], rs["B4"]):
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = center
    filas = [
        ("Total de leads", total),
        ("🔴 Calientes", n_cal),
        ("🟡 Tibios", n_tib),
        ("⚪ Fríos/otros", n_fri),
        ("📈 Derivados a Germán", n_der),
    ]
    for nombre, val in por_plat.items():
        plat_name = {"wa": "WhatsApp", "fb": "Facebook", "ig": "Instagram"}.get(nombre, nombre)
        filas.append((f"Plataforma · {plat_name}", val))
    for nombre, val in filas:
        rs.append([nombre, val])
        r = rs.max_row
        rs.cell(row=r, column=1).border = border
        rs.cell(row=r, column=2).border = border
        rs.cell(row=r, column=2).alignment = center
    rs.column_dimensions["A"].width = 28
    rs.column_dimensions["B"].width = 12

    os.makedirs(OUT_DIR, exist_ok=True)
    nombre_archivo = os.path.join(OUT_DIR, f"Leads_Nico_semana_{fecha_archivo}.xlsx")
    wb.save(nombre_archivo)

    # ── resumen ejecutivo por consola ─────────────────────────────────────────
    print(f"📊 REPORTE SEMANAL DE LEADS — {fecha_pretty} (últimos {DIAS} días)")
    print(f"   Total leads: {total}")
    print(f"   🔴 Calientes: {n_cal} | 🟡 Tibios: {n_tib} | ⚪ Fríos: {n_fri}")
    print(f"   📈 Derivados a Germán: {n_der}")
    print(f"   Por plataforma: " + ", ".join(f"{k}={v}" for k, v in por_plat.items()))
    print(f"   Archivo: {nombre_archivo}")
    print(json.dumps({"archivo": nombre_archivo, "total": total, "calientes": n_cal,
                      "tibios": n_tib, "frios": n_fri, "derivados": n_der}))


if __name__ == "__main__":
    main()
